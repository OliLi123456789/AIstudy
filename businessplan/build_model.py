"""Generate the AIstudy business financial model workbook (v2).

Changes from v1:
  - Chat removed entirely (no chat in the app).
  - AI cost per action is now length-driven: Short/Medium/Long PDFs and
    Slides uploads, token multipliers per action, mix-weighted average.
  - Ads modeled as a Google Ads CPM RANGE ($2.00 - $8.00), with revenue
    and profit computed at both ends.
  - Games are free actions (zero AI cost, still earn ads).

Sheets:
  1. Assumptions   - pricing, doc lengths, upload mix, action multipliers, ads, usage
  2. Cost Per Action - cost matrix by document length + weighted average
  3. Unit Economics - profit per action at $2 and $8 CPM, break-even CPM
  4. Usage Mix      - share of daily actions by type per tier
  5. Per-User P&L   - monthly per tier at both CPM ends
  6. Scale Model    - editable users + 12-month projection
  7. Break-Even     - CPM sweep $2-$8 revenue per action
  8. Scenarios      - heavy-user profit grid: CPM x ad slots

Run:  .venv/bin/python businessplan/build_model.py
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

OUT = "businessplan/aistudy_financial_model.xlsx"

TITLE = Font(bold=True, size=14, color="1F2937")
H1 = Font(bold=True, size=12, color="111827")
H2 = Font(bold=True, size=11, color="374151")
LABEL = Font(size=11)
BOLD = Font(bold=True, size=11)
NOTE = Font(italic=True, size=9, color="6B7280")
INPUT_FILL = PatternFill("solid", fgColor="FEF3C7")
CALC_FILL = PatternFill("solid", fgColor="E0F2FE")
SECTION_FILL = PatternFill("solid", fgColor="E5E7EB")
THIN = Side(style="thin", color="D1D5DB")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def put(ws, row, col, value, font=LABEL, fill=None):
    c = ws.cell(row=row, column=col, value=value)
    c.font = font
    if fill:
        c.fill = fill
    c.border = BOX
    return c


def title(ws, row, text, span):
    put(ws, row, 1, text, TITLE)
    return row + 1


def section(ws, row, text, span):
    for col in range(1, span + 1):
        put(ws, row, col, None, fill=SECTION_FILL)
    put(ws, row, 1, text, H1, SECTION_FILL)
    return row + 1


def hdr(ws, row, headers):
    for col, h in enumerate(headers, 1):
        put(ws, row, col, h, H2, SECTION_FILL)
    return row + 1


def widths(ws, specs):
    for i, w in specs:
        ws.column_dimensions[get_column_letter(i)].width = w


wb = Workbook()

# ============ 1. ASSUMPTIONS ============
A = wb.active
A.title = "Assumptions"
A.sheet_view.showGridLines = False
widths(A, [(46, 46), (22, 22), (44, 44)])

r = 1
r = title(A, r, "AIstudy - Financial Model Assumptions (v2, no chat)", 3)
put(A, r, 1, "Yellow = editable input. Blue = computed. All money USD.", NOTE)
r += 1

r = section(A, r, "DeepSeek V4 Flash pricing (per 1M tokens)", 3)
r = hdr(A, r, ["Parameter", "Value", "Notes"])
P_IN_MISS = r; put(A, r, 1, "Input price (cache miss)"); put(A, r, 2, 0.14, fill=INPUT_FILL); put(A, r, 3, "$ per 1M input tokens")
r += 1
P_IN_HIT = r; put(A, r, 1, "Input price (cache hit)"); put(A, r, 2, 0.0028, fill=INPUT_FILL); put(A, r, 3, "$ per 1M input tokens")
r += 1
P_OUT = r; put(A, r, 1, "Output price"); put(A, r, 2, 0.28, fill=INPUT_FILL); put(A, r, 3, "$ per 1M output tokens")
r += 1
CACHE_HIT = r; put(A, r, 1, "Cache hit rate"); put(A, r, 2, 0.0, fill=INPUT_FILL); put(A, r, 3, "share of input tokens billed at hit price")
r += 1

r = section(A, r, "Document lengths (tokens per upload)", 3)
r = hdr(A, r, ["Length", "Tokens", "Notes"])
LEN_ROWS = {}
lengths = [("Short PDF (~5 pages)", 3300), ("Medium PDF (~20 pages)", 13300),
           ("Long PDF (~60 pages)", 40000), ("Slides deck (~40 slides)", 2000)]
for name, tok in lengths:
    LEN_ROWS[name] = r
    put(A, r, 1, name); put(A, r, 2, tok, fill=INPUT_FILL)
    r += 1

r = section(A, r, "Upload mix (share of uploads, must sum to 1)", 3)
r = hdr(A, r, ["Length", "Share"])
MIX_ROWS = {}
for i, (name, _) in enumerate(lengths):
    share = [0.30, 0.40, 0.20, 0.10][i]
    MIX_ROWS[name] = r
    put(A, r, 1, name); put(A, r, 2, share, fill=INPUT_FILL)
    r += 1

r = section(A, r, "AI action token multipliers", 3)
r = hdr(A, r, ["Action", "Input x doc", "Output x doc", "Notes"])
actions = [
    ("Create note", 1.0, 0.35, "reads full doc, writes condensed notes"),
    ("Generate flashcards", 1.5, 0.9, "doc re-sent for chunking + card text"),
    ("Generate quiz", 2.0, 1.0, "doc + questions + explanations"),
    ("Study All flashcards (folder)", 1.5, 0.9, "per doc, x docs per folder"),
    ("Study All quiz (folder)", 2.0, 1.0, "per doc, x docs per folder"),
    ("Essay grading", 1.5, 0.7, "essay + rubric + feedback"),
]
ACT_ROWS = {}
for name, inm, outm, note in actions:
    ACT_ROWS[name] = r
    put(A, r, 1, name); put(A, r, 2, inm, fill=INPUT_FILL); put(A, r, 3, outm, fill=INPUT_FILL); put(A, r, 4, note)
    r += 1
DOCS_PER_FOLDER = r
put(A, r, 1, "Docs per folder"); put(A, r, 2, 5, fill=INPUT_FILL); put(A, r, 3, "folder actions multiply by this")
r += 1

r = section(A, r, "Lifetime page views per generated asset", 3)
r = hdr(A, r, ["Action", "Views", "Notes"])
views_vals = [12, 15, 4, 15, 4, 2]
views_notes = [
    "reopened while studying",
    "drilled repeatedly before exams",
    "retake attempts",
    "folder drills",
    "folder retakes",
    "grade + one review",
]
VIEW_ROWS = []
for i, ((name, _, _, _), v, note) in enumerate(zip(actions, views_vals, views_notes)):
    VIEW_ROWS.append(r)
    put(A, r, 1, name); put(A, r, 2, v, fill=INPUT_FILL); put(A, r, 3, note)
    r += 1

r = section(A, r, "Ad revenue (Google Ads, CPM range $2.00 - $8.00)", 3)
r = hdr(A, r, ["Parameter", "Value", "Notes"])
AD_SLOTS = r; put(A, r, 1, "Ad slots per page"); put(A, r, 2, 2, fill=INPUT_FILL); put(A, r, 3, "no rotation: 1 action = 1 page view")
r += 1
CPM_LOW = r; put(A, r, 1, "CPM low ($)"); put(A, r, 2, 2.00, fill=INPUT_FILL)
r += 1
CPM_HIGH = r; put(A, r, 1, "CPM high ($)"); put(A, r, 2, 8.00, fill=INPUT_FILL)
r += 1
VIEW_STD = r; put(A, r, 1, "Viewability standard"); put(A, r, 2, 0.60, fill=INPUT_FILL)
r += 1
VIEW_STICKY = r; put(A, r, 1, "Viewability sticky"); put(A, r, 2, 0.95, fill=INPUT_FILL)
r += 1
ADBLOCK = r; put(A, r, 1, "Ad block rate"); put(A, r, 2, 0.30, fill=INPUT_FILL)
r += 1
USE_STICKY = r; put(A, r, 1, "Use sticky viewability? (TRUE/FALSE)"); put(A, r, 2, True, fill=INPUT_FILL)
r += 1

r = section(A, r, "Hosting & infrastructure", 3)
r = hdr(A, r, ["Parameter", "Value", "Notes"])
HOST_FIXED = r; put(A, r, 1, "Fixed hosting ($/mo)"); put(A, r, 2, 20, fill=INPUT_FILL); put(A, r, 3, "CDN + app server + DB base")
r += 1
HOST_PER_MAU = r; put(A, r, 1, "Variable hosting ($/MAU/mo)"); put(A, r, 2, 0.005, fill=INPUT_FILL); put(A, r, 3, "storage + bandwidth + DB per active user")
r += 1

r = section(A, r, "Usage per tier (chat removed)", 4)
r = hdr(A, r, ["Tier", "Actions/day", "Days/month", "Free actions/day (games)"])
tiers = [
    ("Heavy (3+ courses)", 12, 25, 3),
    ("Moderate (2-3 courses)", 5, 20, 2),
    ("Light (1-2 courses)", 2, 15, 1),
    ("Casual", 0.5, 10, 0.5),
]
TIER_ROWS = []
for name, ad, dm, free in tiers:
    TIER_ROWS.append(r)
    put(A, r, 1, name); put(A, r, 2, ad, fill=INPUT_FILL); put(A, r, 3, dm, fill=INPUT_FILL); put(A, r, 4, free, fill=INPUT_FILL)
    r += 1

r = section(A, r, "Computed", 3)
r = hdr(A, r, ["Metric", "Value", "Formula"])
EFF_IN = r
put(A, r, 1, "Effective input price ($/1M)")
put(A, r, 2, f"=B{P_IN_MISS}*(1-B{CACHE_HIT})+B{P_IN_HIT}*B{CACHE_HIT}", fill=CALC_FILL)
put(A, r, 3, "blend of cache miss/hit")
r += 1
REV_LOW = r
put(A, r, 1, "Revenue per PAGE VIEW at CPM low")
put(A, r, 2, f"=IF(B{USE_STICKY},B{CPM_LOW}*B{VIEW_STICKY},B{CPM_LOW}*B{VIEW_STD})*(1-B{ADBLOCK})*B{AD_SLOTS}/1000", fill=CALC_FILL)
r += 1
REV_HIGH = r
put(A, r, 1, "Revenue per PAGE VIEW at CPM high")
put(A, r, 2, f"=IF(B{USE_STICKY},B{CPM_HIGH}*B{VIEW_STICKY},B{CPM_HIGH}*B{VIEW_STD})*(1-B{ADBLOCK})*B{AD_SLOTS}/1000", fill=CALC_FILL)
r += 1

# ============ 2. COST PER ACTION ============
C = wb.create_sheet("Cost Per Action")
C.sheet_view.showGridLines = False
widths(C, [(28, 28), (12, 12), (12, 12), (12, 12), (12, 12), (16, 16)])

r = 1
r = title(C, r, "AI cost per action by document length", 6)
put(C, r, 1, "Cost = (in-mult x doc-tokens x eff-input-price + out-mult x doc-tokens x output-price) / 1M. Folder actions x docs per folder.", NOTE)
r += 1
r = hdr(C, r, ["Action", "Short", "Medium", "Long", "Slides", "Weighted avg"])
C_START = r
len_names = [n for n, _ in lengths]
for i, (name, _, _, _) in enumerate(actions):
    rr = C_START + i
    arow = ACT_ROWS[name]
    put(C, rr, 1, name)
    for j, lname in enumerate(len_names):
        lrow = LEN_ROWS[lname]
        mrow = MIX_ROWS[lname]
        if "folder" in name:
            formula = (f"=(Assumptions!B{arow}*Assumptions!B{lrow}*Assumptions!B{EFF_IN}"
                       f"+Assumptions!C{arow}*Assumptions!B{lrow}*Assumptions!B{P_OUT})"
                       f"/1000000*Assumptions!B{DOCS_PER_FOLDER}")
        else:
            formula = (f"=(Assumptions!B{arow}*Assumptions!B{lrow}*Assumptions!B{EFF_IN}"
                       f"+Assumptions!C{arow}*Assumptions!B{lrow}*Assumptions!B{P_OUT})/1000000")
        put(C, rr, 2 + j, formula, fill=CALC_FILL)
    mix_terms = "+".join(f"{col}{rr}*Assumptions!B{MIX_ROWS[n]}"
                         for col, n in zip("BCDE", len_names))
    put(C, rr, 6, f"={mix_terms}", BOLD, CALC_FILL)
C_END = C_START + len(actions) - 1

# ============ 3. UNIT ECONOMICS ============
U = wb.create_sheet("Unit Economics")
U.sheet_view.showGridLines = False
widths(U, [(28, 28), (12, 12), (14, 14), (14, 14), (14, 14), (14, 14), (18, 18)])

r = 1
r = title(U, r, "Unit economics per action (CPM $2 vs $8)", 7)
put(U, r, 1, "Weighted avg cost across upload mix. Revenue per action is flat across actions (1:1 action-to-ad).", NOTE)
r += 1
r = hdr(U, r, ["Action", "Avg cost", "Revenue @$2", "Revenue @$8", "Profit @$2", "Profit @$8", "Break-even CPM ($)"])
U_START = r
for i, (name, _, _, _) in enumerate(actions):
    rr = U_START + i
    vrow = VIEW_ROWS[i]
    put(U, rr, 1, name)
    put(U, rr, 2, f"='Cost Per Action'!F{C_START + i}", fill=CALC_FILL)
    put(U, rr, 3, f"=Assumptions!B{vrow}*Assumptions!B{REV_LOW}", fill=CALC_FILL)
    put(U, rr, 4, f"=Assumptions!B{vrow}*Assumptions!B{REV_HIGH}", fill=CALC_FILL)
    put(U, rr, 5, f"=C{rr}-B{rr}", BOLD, CALC_FILL)
    put(U, rr, 6, f"=D{rr}-B{rr}", BOLD, CALC_FILL)
    put(U, rr, 7, (f"=B{rr}*1000/(Assumptions!B{vrow}*Assumptions!B{AD_SLOTS}"
                   f"*IF(Assumptions!B{USE_STICKY},Assumptions!B{VIEW_STICKY},Assumptions!B{VIEW_STD})"
                   f"*(1-Assumptions!B{ADBLOCK}))"), fill=CALC_FILL)
U_END = U_START + len(actions) - 1

# ============ 4. USAGE MIX ============
M = wb.create_sheet("Usage Mix")
M.sheet_view.showGridLines = False
widths(M, [(28, 28), (12, 12), (12, 12), (12, 12)])

r = 1
r = title(M, r, "Action mix - share of daily paid actions by tier", 4)
r += 1
r = hdr(M, r, ["Action", "Heavy", "Moderate", "Light/Casual"])
mix = [
    ("Create note", 0.10, 0.10, 0.15),
    ("Generate flashcards", 0.15, 0.15, 0.20),
    ("Generate quiz", 0.10, 0.10, 0.10),
    ("Study All flashcards", 0.20, 0.20, 0.15),
    ("Study All quiz", 0.15, 0.15, 0.10),
    ("Essay grading", 0.10, 0.10, 0.10),
]
M_START = r
for name, h, m_, l in mix:
    put(M, r, 1, name); put(M, r, 2, h, fill=INPUT_FILL); put(M, r, 3, m_, fill=INPUT_FILL); put(M, r, 4, l, fill=INPUT_FILL)
    r += 1
M_END = r - 1
put(M, r, 1, "Total paid share", H1, SECTION_FILL)
put(M, r, 2, f"=SUM(B{M_START}:B{M_END})", fill=CALC_FILL)
put(M, r, 3, f"=SUM(C{M_START}:C{M_END})", fill=CALC_FILL)
put(M, r, 4, f"=SUM(D{M_START}:D{M_END})", fill=CALC_FILL)

# ============ 5. PER-USER P&L ============
P = wb.create_sheet("Per-User P&L")
P.sheet_view.showGridLines = False
widths(P, [(28, 28), (14, 14), (14, 14), (12, 12), (14, 14), (14, 14), (14, 14), (14, 14)])

r = 1
r = title(P, r, "Monthly P&L per user tier (paid + free actions)", 8)
r += 1
r = hdr(P, r, ["Tier", "Paid acts/mo", "Free acts/mo", "AI cost", "Revenue @$2", "Profit @$2", "Revenue @$8", "Profit @$8"])
P_START = r
mix_cols = ["B", "C", "D", "D"]
for i, trow in enumerate(TIER_ROWS):
    rr = P_START + i
    put(P, rr, 1, tiers[i][0])
    put(P, rr, 2, f"=Assumptions!B{trow}*Assumptions!C{trow}", fill=CALC_FILL)
    put(P, rr, 3, f"=Assumptions!D{trow}*Assumptions!C{trow}", fill=CALC_FILL)
    cost_terms = "+".join(
        f"'Usage Mix'!{mix_cols[i]}{M_START + j}*'Cost Per Action'!F{C_START + j}"
        for j in range(len(actions))
    )
    put(P, rr, 4, f"=B{rr}*({cost_terms})", fill=CALC_FILL)
    views_terms = "+".join(
        f"'Usage Mix'!{mix_cols[i]}{M_START + j}*Assumptions!B{VIEW_ROWS[j]}"
        for j in range(len(actions))
    )
    put(P, rr, 5, f"=B{rr}*({views_terms})*Assumptions!B{REV_LOW}+C{rr}*Assumptions!B{REV_LOW}", fill=CALC_FILL)
    put(P, rr, 6, f"=E{rr}-D{rr}", BOLD, CALC_FILL)
    put(P, rr, 7, f"=B{rr}*({views_terms})*Assumptions!B{REV_HIGH}+C{rr}*Assumptions!B{REV_HIGH}", fill=CALC_FILL)
    put(P, rr, 8, f"=G{rr}-D{rr}", BOLD, CALC_FILL)
P_END = P_START + len(TIER_ROWS) - 1

# ============ 6. SCALE MODEL ============
S = wb.create_sheet("Scale Model")
S.sheet_view.showGridLines = False
widths(S, [(22, 22), (12, 12), (12, 12), (12, 12), (12, 12), (14, 14), (14, 14), (14, 14), (14, 14)])

r = 1
r = title(S, r, "Scale model - total business P&L", 9)
put(S, r, 1, "Profit = ad revenue - AI cost - hosting & infra.", NOTE)
r += 1
r = hdr(S, r, ["", "Heavy", "Moderate", "Light", "Casual", "AI cost", "Rev @$2", "Profit @$2", "Rev @$8", "Profit @$8"])
r += 1
r = hdr(S, r, ["Users (editable)", 500, 1500, 4000, 10000, "", "", "", "", ""])
USER_ROW = r - 1
for col in (2, 3, 4, 5):
    S.cell(row=USER_ROW, column=col).fill = INPUT_FILL
    S.cell(row=USER_ROW, column=col).border = BOX
    S.cell(row=USER_ROW, column=col).font = BOLD
r += 1
def sumprod4(col, row):
    """Explicit per-tier sum: Per-User P&L col x B..E user cells at `row`."""
    return "=" + "+".join(
        f"'Per-User P&L'!{col}{P_START + i}*{letter}{row}"
        for i, letter in enumerate("BCDE")
    )


metrics = [
    ("AI cost", sumprod4("D", USER_ROW)),
    ("Hosting", f"=Assumptions!B{HOST_FIXED}+Assumptions!B{HOST_PER_MAU}*SUM(B{USER_ROW}:E{USER_ROW})"),
    ("Revenue @$2", sumprod4("E", USER_ROW)),
    ("Profit @$2", None),
    ("Revenue @$8", sumprod4("G", USER_ROW)),
    ("Profit @$8", None),
]
METRIC_ROWS = {}
for name, formula in metrics:
    METRIC_ROWS[name] = r
    put(S, r, 1, name, H2)
    put(S, r, 6, formula if formula else "", BOLD, CALC_FILL)
    r += 1
put(S, METRIC_ROWS["Profit @$2"], 6,
    f"=F{METRIC_ROWS['Revenue @$2']}-F{METRIC_ROWS['AI cost']}-F{METRIC_ROWS['Hosting']}", BOLD, CALC_FILL)
put(S, METRIC_ROWS["Profit @$8"], 6,
    f"=F{METRIC_ROWS['Revenue @$8']}-F{METRIC_ROWS['AI cost']}-F{METRIC_ROWS['Hosting']}", BOLD, CALC_FILL)
put(S, r, 1, "Total users")
put(S, r, 6, f"=SUM(B{USER_ROW}:E{USER_ROW})", BOLD, CALC_FILL)
r += 2

r = section(S, r, "12-month projection (5% monthly user growth)", 9)
r = hdr(S, r, ["Month", "Heavy", "Moderate", "Light", "Casual", "Net profit @$2", "Net profit @$8", "Rev @$8", ""])
PROJ_START = r
for m in range(12):
    rr = PROJ_START + m
    put(S, rr, 1, f"Month {m+1}")
    for col, letter in ((2, "B"), (3, "C"), (4, "D"), (5, "E")):
        if m == 0:
            put(S, rr, col, f"={letter}{USER_ROW}", fill=CALC_FILL)
        else:
            put(S, rr, col, f"={letter}{USER_ROW}*(1+0.05)^{m}", fill=CALC_FILL)
    put(S, rr, 6, f"={sumprod4('E', rr)[1:]}-{sumprod4('D', rr)[1:]}-(Assumptions!B{HOST_FIXED}+Assumptions!B{HOST_PER_MAU}*SUM(B{rr}:E{rr}))", fill=CALC_FILL)
    put(S, rr, 7, f"={sumprod4('G', rr)[1:]}-{sumprod4('D', rr)[1:]}-(Assumptions!B{HOST_FIXED}+Assumptions!B{HOST_PER_MAU}*SUM(B{rr}:E{rr}))", fill=CALC_FILL)
    put(S, rr, 8, sumprod4("G", rr), fill=CALC_FILL)
tot = PROJ_START + 12
put(S, tot, 1, "Year 1 total", H1, SECTION_FILL)
put(S, tot, 6, f"=SUM(F{PROJ_START}:F{PROJ_START+11})", H1, CALC_FILL)
put(S, tot, 7, f"=SUM(G{PROJ_START}:G{PROJ_START+11})", H1, CALC_FILL)
put(S, tot, 8, f"=SUM(H{PROJ_START}:H{PROJ_START+11})", H1, CALC_FILL)

# ============ 7. BREAK-EVEN ============
B = wb.create_sheet("Break-Even")
B.sheet_view.showGridLines = False
widths(B, [(42, 42), (20, 20), (20, 20)])

r = 1
r = title(B, r, "Break-even analysis", 3)
r += 1
r = section(B, r, "Revenue per page view across Google Ads CPM range", 3)
r = hdr(B, r, ["CPM ($)", "Rev/view sticky (95%)", "Rev/view standard (60%)"])
b_start = r
for i, cpm in enumerate([2.0, 3.0, 4.0, 5.0, 6.0, 7.0, 8.0]):
    rr = b_start + i
    put(B, rr, 1, cpm, fill=INPUT_FILL)
    put(B, rr, 2, (f"=A{rr}*Assumptions!B{AD_SLOTS}*Assumptions!B{VIEW_STICKY}"
                   f"*(1-Assumptions!B{ADBLOCK})/1000"), fill=CALC_FILL)
    put(B, rr, 3, (f"=A{rr}*Assumptions!B{AD_SLOTS}*Assumptions!B{VIEW_STD}"
                   f"*(1-Assumptions!B{ADBLOCK})/1000"), fill=CALC_FILL)
r = b_start + 8
r = section(B, r, "Break-even CPM per action (2 slots, sticky)", 3)
r = hdr(B, r, ["Action", "Avg cost", "Break-even CPM ($)"])
b2 = r
for i in range(len(actions)):
    rr = b2 + i
    put(B, rr, 1, actions[i][0])
    put(B, rr, 2, f"='Cost Per Action'!F{C_START + i}", fill=CALC_FILL)
    put(B, rr, 3, f"='Unit Economics'!G{U_START + i}", BOLD, CALC_FILL)

# ============ 8. SCENARIOS ============
S2 = wb.create_sheet("Scenarios")
S2.sheet_view.showGridLines = False
widths(S2, [(34, 34), (16, 16), (16, 16), (16, 16)])

r = 1
r = title(S2, r, "Heavy user monthly profit: CPM x ad slots", 4)
put(S2, r, 1, "Heavy user = 300 paid (x lifetime views) + 75 free game views/month. AI cost from Per-User P&L.", NOTE)
r += 1
r = hdr(S2, r, ["CPM ($)", "1 slot", "2 slots", "3 slots"])
s2_views = "+".join(f"'Usage Mix'!B{M_START + j}*Assumptions!B{VIEW_ROWS[j]}"
                    for j in range(len(actions)))
for i, cpm in enumerate([2.0, 4.0, 6.0, 8.0]):
    rr = r + i
    put(S2, rr, 1, cpm, fill=INPUT_FILL)
    for j, slots in enumerate([1, 2, 3]):
        rev = (f"=('Per-User P&L'!B{P_START}*({s2_views})+'Per-User P&L'!C{P_START})"
               f"*A{rr}*{slots}"
               f"*IF(Assumptions!B{USE_STICKY},Assumptions!B{VIEW_STICKY},Assumptions!B{VIEW_STD})"
               f"*(1-Assumptions!B{ADBLOCK})/1000-'Per-User P&L'!D{P_START}")
        put(S2, rr, 2 + j, rev, BOLD, CALC_FILL)

wb.save(OUT)
print("Saved", OUT)
print("Sheets:", wb.sheetnames)
